from io import BytesIO

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.schemas import ImportData, GPSImportData
from app.services import (
    import_to_db,
    import_gps_objects,
    import_service_connections,
    normalize_row,
)

router = APIRouter(prefix="/import", tags=["Import"])


def import_response(result):
    return {
        "status": "success",
        "imported_count": len(result["imported"]),
        "errors_count": len(result["errors"]),
        "warnings_count": len(result["warnings"]),
        **{key: value for key, value in result.items() if key != "imported"},
    }


@router.post("/excel")
async def import_excel(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")

    content = await file.read()
    workbook = load_workbook(filename=BytesIO(content), data_only=True)

    if "Modified" not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail="Sheet 'Modified' not found")
    worksheet = workbook["Modified"]

    normalized_rows = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
        normalized_rows.extend({**entry, "_excel_row": row_number} for entry in normalize_row(row))
    workbook.close()

    import_result = await import_service_connections(normalized_rows, db)

    return import_response(import_result)


@router.post("/service-connections")
async def import_service_connections_endpoint(payload: list[dict], db: AsyncSession = Depends(get_db)):
    if not payload:
        raise HTTPException(status_code=400, detail="Empty payload")

    import_result = await import_service_connections(payload, db)
    return import_response(import_result)


@router.post("/gps")
async def import_gps(payload: GPSImportData, db: AsyncSession = Depends(get_db)):
    if not payload.points:
        raise HTTPException(status_code=400, detail="Empty payload")

    result = await import_gps_objects(payload, db, return_report=True)
    return {**import_response(result), "imported": len(result["imported"])}


@router.post("/gps-legacy")
async def import_gps_legacy(payload: list[dict], db: AsyncSession = Depends(get_db)):
    if not payload:
        raise HTTPException(status_code=400, detail="Empty payload")

    import_data = ImportData(raw_entries=payload)
    result = await import_to_db(import_data)
    return {
        **import_response(result),
        "object_stats": result["objects"],
        "objects": len(import_data.objects),
        "service_connections": len(import_data.service_connections),
    }
