from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from io import BytesIO

from frs_energy_data.schemas import ImportData, GPSImportData
from frs_energy_data.services import import_to_db, import_gps_objects, import_feeders_objects, normalize_row
from frs_energy_data.database import get_db

router = APIRouter(prefix="/import", tags=["Import"])

@router.post("/excel")
async def import_excel(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")

    # Datei in BytesIO laden
    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)

    # Arbeitsblatt auswählen (hier "Modified")
    if "Modified" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Sheet 'Modified' not found")
    ws = wb["Modified"]

    # Excel-Zeilen normalisieren
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        parsed = normalize_row(row)
        result.extend(parsed)

    # Dann die Import-Logik aufrufen
    import_result = await import_feeders_objects(result, db)

    return {
    "status": "success",
    "imported_count": len(import_result.get("imported", [])),
    "errors_count": len(import_result.get("errors", [])),
    "errors": import_result.get("errors", [])
}

@router.post("/feeders")
async def import_feeders(payload: list[dict], db: AsyncSession = Depends(get_db)):
    if not payload:
        raise HTTPException(status_code=400, detail="Empty payload")

    import_result = await import_feeders_objects(payload, db)
    return {
    "status": "success",
    "imported_count": len(import_result.get("imported", [])),
    "errors_count": len(import_result.get("errors", [])),
    "errors": import_result.get("errors", [])
}

@router.post("/gps")
async def import_gps(payload: GPSImportData, db: AsyncSession = Depends(get_db)):
    """
    Importiert nur GPS-Points (Objects) in die Datenbank.
    """
    if not payload.points:
        raise HTTPException(status_code=400, detail="Empty payload")

    imported_objects = await import_gps_objects(payload, db)
    return {"status": "success", "imported": len(imported_objects)}

@router.post("/gps_old")
async def import_gps(payload: list[dict], db: AsyncSession = Depends(get_db)):
    """
    Importiere JSON Payload direkt in die DB.
    """
    if not payload:
        raise HTTPException(status_code=400, detail="Empty payload")

    import_data = ImportData(raw_entries=payload)
    await import_to_db(import_data)
    return {"status": "success", "objects": len(import_data.objects), "feeders": len(import_data.feeders)}